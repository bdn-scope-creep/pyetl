from openpyxl import load_workbook
import sys
import time
import os
import re
import shutil
import requests
from itertools import islice
from mysql.connector import connect, Error

d_filename = sys.argv[1]
p_filename = sys.argv[2]

connection = connect(host=os.environ['SCOPECREEP_DB_HOST'],
        port=32939,
        user='admin',
        password=os.environ['SCOPECREEP_DB_PASSWORD'],
        database='DATA')

# Name
# Title_URL
# Image
# Tier
# Partner Type
# Geo
d_wb = load_workbook(d_filename, read_only=True)
d_sheet = d_wb.active


# Address
# Description
# Verticals
# Geos
# Image_URL
# Page_URL
p_wb = load_workbook(p_filename, read_only=True)
p_sheet = p_wb.active

SERVICE_PROVIDER_SELECT_SQL = "SELECT {0} FROM {1} WHERE {2} = '{3}'"
GEO_SELECT_SQL = "SELECT {0} FROM {1} WHERE {2} = '{3}'"
GEO_INSERT_SQL = "INSERT INTO {0} ({1}) VALUES('{2}')"
VERTICAL_SELECT_SQL = "SELECT {0} FROM {1} WHERE {2} = '{3}'"
VERTICAL_INSERT_SQL = "INSERT INTO {0} ({1}) VALUES('{2}')"
PARTNERSHIP_SELECT_SQL= "SELECT {0} FROM {1} WHERE year = 2022 AND category_id = 1 AND {2} = '{3}'"
PARTNERSHIP_INSERT_SQL = "INSERT INTO {0} ({1},year,category_id) VALUES('{2}',2022,1)"


SELECT_SQL = {
        'geo': GEO_SELECT_SQL,
        'vertical': VERTICAL_SELECT_SQL,
        'partnership': PARTNERSHIP_SELECT_SQL,
        'service_provider':SERVICE_PROVIDER_SELECT_SQL
        }
INSERT_SQL = {
        'geo': GEO_INSERT_SQL,
        'vertical': VERTICAL_INSERT_SQL,
        'partnership': PARTNERSHIP_INSERT_SQL
        }

def find_id(table, id_col, col, val):
    with connection.cursor() as cursor:
        s = SELECT_SQL[table].format(id_col, table, col, val)
        cursor.execute(s)
        row = cursor.fetchone()
        if not row: return None
        return row[0]


def find_or_insert(table, id_col, col, val):
    id = find_id(table, id_col, col, val)
    if id is None:
        with connection.cursor() as cursor:
            s = INSERT_SQL[table].format(table, col, val)
            cursor.execute(s)
            connection.commit()
            return cursor.lastrowid
    return id

def find_matching_profile(profile_url):
    for row in p_sheet.rows:
        page_url = row[5].value.strip()
        if profile_url == page_url:
            return row
    return None

def parse_address(address):
    address = address[12:] 
    result = []
    for line in address.split("\n"):
        line = line.strip()
        if line.startswith("Headquar") or line.startswith("Phone"):
            continue
        if line:
            result += [line]
    return "\n".join(result)

def parse_verticals(verticals):
    result = []
    if verticals is None:
        return result
    verticals = verticals.replace('<li>', '').replace('</li>', '\n').replace('&amp;', '&')
    for line in verticals.split('\n'):
        line = line.strip()
        if line:
            result += [line]
    return result

def download_image(name, image_url):
    name = unique_name(name)
    filename = name + '-logo300x150.png'
    localfilename = 'images/' + filename
    s3filename = "https://sc-image-94c553e8.s3.amazonaws.com/service_provider/" + filename
    if os.path.exists(localfilename): return s3filename
    r = requests.get(image_url, stream = True)
    if r.status_code == 200:
        # Set decode_content value to True, otherwise the downloaded image file's size will be zero.
        r.raw.decode_content = True
        with open(localfilename,'wb') as f:
            shutil.copyfileobj(r.raw, f)
        time.sleep(0.5)
        return s3filename
    return None

def insert(table, d):
    with connection.cursor() as cursor:
        placeholder = ", ".join(["%s"] * len(d))
        stmt = "insert into `{table}` ({columns}) values ({values});".format(table=table, columns=",".join(d.keys()), values=placeholder)
        cursor.execute(stmt, list(d.values()))
        connection.commit()
        return cursor.lastrowid

def unique_name(name):
    name = re.sub('[^0-9a-zA-Z]+', '_', name.strip()).replace('__', '_')
    name = re.sub('^_', '', name)
    name = re.sub('_$', '', name)
    return name

def find_or_insert_service_provider(d):
    table = 'service_provider'
    id_col = 'service_provider_id'
    col = 'unique_name'
    val = d['unique_name']
    id = find_id(table, id_col, col, val)
    if id is None:
        return insert(table, d)
    return id


for row in islice(d_sheet.rows, 1, None):
    name = row[0].value.strip()
    profile_url = row[1].value.strip()
    image_url = row[2].value.strip()
    partnership = row[3].value.split(': ')[1].strip()
    if 'registered' in partnership.lower().strip(): continue
    partner_type = row[4].value.split(': ')[1].strip()
    if not ('gsi' in partner_type.lower() or 'rsi' in partner_type.lower()): continue
    profile = find_matching_profile(profile_url)
    if profile is None:
        print("Could not find profile for ", name)
        sys.exit(0)
    address = parse_address(profile[0].value)
    description =  profile[1].value
    verticals = parse_verticals(profile[2].value)
    image_url =  profile[4].value
    try:

        primary_geo = row[5].value.split(': ')[1].strip()
    except IndexError:
        continue
    logo_url = None
    if image_url:
        logo_url = download_image(name, image_url)
    d = {
            "name": name,
            "unique_name": unique_name(name).lower(),
            "logo_url": logo_url,
            "source_url": profile_url,
            "description": description,
            "published": "TRUE"
        }
    service_provider_id = find_or_insert_service_provider(d)
    print("service_provider_id = ", service_provider_id)
    partnership_id = find_or_insert('partnership','partnership_id','name', partnership)
    # CATEGORY
    insert('service_provider_category', {
        'service_provider_id': service_provider_id,
        'category_id':1,
        'source':profile_url
        })
    # PARTNERSHIP
    insert('service_provider_partnership', {
        'service_provider_id': service_provider_id,
        'partnership_id':partnership_id
        })
    # GEO
    for geo in primary_geo.split(';'):
        geo = geo.strip()
        geo_id = find_or_insert('geo','geo_id','name', geo)
        insert('service_provider_geo', {
            'service_provider_id': service_provider_id,
            'geo_id':geo_id
            })
    # VERTICAL
    for vertical in verticals:
        vertical_id = find_or_insert('vertical','vertical_id','name', vertical)
        insert('service_provider_vertical', {
            'service_provider_id': service_provider_id,
            'vertical_id':vertical_id
            })
